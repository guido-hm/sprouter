import tkinter as tk
import tkinter.font
import openpyxl
import datetime
import time as tm
import os




def read_excel():
	#date = datetime.datetime()
	#jar_backup = openpyxl.load_workbook('sprouts_tracker.xlsx')   ### These three commented lines make a backup of the file each time before I edit it. Need to give it a location on where to save backups
	#jar_backup.save('sprouts_backup_%s' %s str(date))
	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)

	sheet = jar_database['Database']
	sheet2 = jar_database['Cost']
	jar_list = []

	ROW = 2
	COLUMN = 2

	cell = sheet.cell(row=ROW, column = COLUMN).value
	while bool(cell) == True:
		jar = [str(cell)] # Resets value of jar to contain next jar date
		date_of_current_cell = cell.split('/')
		for i in range(len(date_of_current_cell)): # These next few lines of code get the date jar was started and todays date. If the difference between the days is >= the 'Total Days' a jar is supposed to be spent growing AND the current status == 'growing', then it will chane the status to 'ready' and Entry Color to 'yellow'
			date_of_current_cell[i] = int(date_of_current_cell[i])
		if ((datetime.date.today() - datetime.date(date_of_current_cell[2], date_of_current_cell[0], date_of_current_cell[1])).days) >= int(sheet.cell(row=10, column=COLUMN).value) and (sheet.cell(row=3, column=COLUMN).value == 'growing'):
			sheet.cell(row=3, column=COLUMN, value='ready') #If jar met the days it was goning to grow for, this updates the status from 'growing' to 'ready'
			sheet.cell(row=4, column=COLUMN, value='#FDFD98') #If jar met the days it was goning to grow for, this updates the Entry Color from 'green' to 'yellow'
			jar_database.save(file_rel_path) 

		ROW = 3
		cell = sheet.cell(row=ROW, column = COLUMN).value

		while cell is not None:
			jar.append(str(cell)) #This appends to jar AS A STRING!!!
			ROW += 1
			cell = sheet.cell(row=ROW, column = COLUMN).value
			print('CELL: ', cell)
			print('CELL: ', type(cell))


		ROW = 2
		COLUMN += 1
		cell = sheet.cell(row=ROW, column = COLUMN).value
		jar_list.append(jar) # Adds jar to jarlist


	jar_list_flipped = []
	for i in range(1, len(jar_list)+1):
		temp_var = jar_list[-i]
		jar_list_flipped.append(temp_var)

	num_of_jars = COLUMN - 2

	# Gets the current water and seed prices
	current_water_cost = sheet2.cell(row=2, column=2).value
	#current_seed_cost = sheet2.cell(row=5, column=2).value SHOULD PROBABLY DELETE THIS LINE IF CODE WORKS AFTER TEST RUN 6/25/2020


	return jar_list_flipped, num_of_jars, current_water_cost


	#this will most likely have to be called at the beginning of the program (or before the buttons) to be able to fill in the buttons with proper info.

def move_left():
	global position
	global button_left
	global button_right
	print(position, 'POSS')
	position += 1



	button1 = tk.Button(frame_1, bg=jar_list[position+4][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+4][0],jar_list[position+4][1],(datetime.date.today()-jar_dates[position+4]).days,jar_list[position+4][8]), command=lambda: show_info(jar_list[position+4], position+4)) # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg=jar_list[position+3][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+3][0],jar_list[position+3][1],(datetime.date.today()-jar_dates[position+3]).days,jar_list[position+3][8]), command=lambda: show_info(jar_list[position+3], position+3))
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)

	if position == len(jar_list)-5:
		button_left = tk.Button(frame_left, bg='gray', text='<<', command=move_left, state='disabled') # Disables the button 
		button_left.place(relwidth=1, relheight=1)

	button_right = tk.Button(frame_right, bg='#FF7000', text='>>', command=move_right, state='normal') # Enables the button
	button_right.place(relwidth=1, relheight=1)

def move_right():

	global position
	global button_left
	global button_right


	position -= 1
	button1 = tk.Button(frame_1, bg=jar_list[position+4][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+4][0],jar_list[position+4][1],(datetime.date.today()-jar_dates[position+4]).days,jar_list[position+4][8]), command=lambda: show_info(jar_list[position+4], position+4)) # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg=jar_list[position+3][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+3][0],jar_list[position+3][1],(datetime.date.today()-jar_dates[position+3]).days,jar_list[position+3][8]), command=lambda: show_info(jar_list[position+3], position+3))
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)

	if position == 0:
		button_right = tk.Button(frame_right, bg='gray', text='>>', command=move_right, state='disabled')
		button_right.place(relwidth=1, relheight=1)

	button_left = tk.Button(frame_left, bg='#FF7000', text='<<', command=move_left, state='normal') # Disables the button 
	button_left.place(relwidth=1, relheight=1)

def type_to_excel(date, seed_amt, weight, floating_seeds, rinses_per_day, days_in_dark, total_days, initial_rinse, new_column, window):

	global jar_list
	global num_of_jars
	global jar_dates
	global position

	#OPENS EXCEL NOTEBOOK
	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)
	sheet = jar_database['Database']
	sheet2= jar_database['Cost']
	sheet.cell(row=2, column=new_column, value=date) # Writes the Date
	sheet.cell(row=3, column=new_column, value='growing')
	sheet.cell(row=4, column=new_column, value='#85DE77')
	sheet.cell(row=5, column=new_column, value=float(seed_amt))
	sheet.cell(row=6, column=new_column, value=float(weight))
	sheet.cell(row=7, column=new_column, value=int(floating_seeds))
	sheet.cell(row=8, column=new_column, value=int(rinses_per_day))
	sheet.cell(row=9, column=new_column, value=int(days_in_dark))
	sheet.cell(row=10, column=new_column, value=int(total_days))
	sheet.cell(row=11, column=new_column, value=float(initial_rinse))
	sheet.cell(row=12, column=new_column, value='N/A')
	sheet.cell(row=13, column=new_column, value='N/A')
	sheet.cell(row=14, column=new_column, value=sheet2.cell(row=2, column=2).value)
	sheet.cell(row=15, column=new_column, value=sheet2.cell(row=5, column=position_seed_provider+2).value)
	sheet.cell(row=16, column=new_column, value='N/A')
	sheet.cell(row=17, column=new_column, value='N/A')
	sheet.cell(row=18, column=new_column, value='N/A')
	sheet.cell(row=19, column=new_column, value=list_seed_providers[position_seed_provider])

	jar_database.save(file_rel_path)

	position = 0

	read_excel_var = read_excel() # This returns a list of the jars and the integer of amount of jars

	jar_list = read_excel_var[0] # This separates the list from read_excel_var to its own variable

	num_of_jars = read_excel_var[1] # num_of_jars is updated here when a Jar is added.





	jar_dates = []
	for i in range(len(jar_list)):
		date = jar_list[i][0].split('/')
		for j in range(len(date)):
			date[j] = int(date[j])

		date = datetime.date(date[2], date[0], date[1])
		jar_dates.append(date)


	if num_of_jars >= 5:
		button1 = tk.Button(frame_1, bg=jar_list[position+4][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+4][0],jar_list[position+4][1],(datetime.date.today()-jar_dates[position+4]).days,jar_list[position+4][8]), command=lambda: show_info(jar_list[position+4], position+4)) # Leftmost button
		button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

		button2 = tk.Button(frame_1, bg=jar_list[position+3][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+3][0],jar_list[position+3][1],(datetime.date.today()-jar_dates[position+3]).days,jar_list[position+3][8]), command=lambda: show_info(jar_list[position+3], position+3))
		button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

		button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
		button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

		button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
		button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

		button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
		button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
	if num_of_jars == 4:
		button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
		button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

		button2 = tk.Button(frame_1, bg=jar_list[position+3][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+3][0],jar_list[position+3][1],(datetime.date.today()-jar_dates[position+3]).days,jar_list[position+3][8]), command=lambda: show_info(jar_list[position+3], position+3))
		button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

		button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
		button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

		button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
		button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

		button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
		button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
	if num_of_jars == 3:
		button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
		button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

		button2 = tk.Button(frame_1, bg='gray', state='disabled')
		button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

		button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
		button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

		button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
		button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

		button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
		button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
	if num_of_jars == 2:
		button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
		button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

		button2 = tk.Button(frame_1, bg='gray', state='disabled')
		button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

		button3 = tk.Button(frame_1, bg='gray', state='disabled')
		button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

		button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
		button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

		button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
		button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
	if num_of_jars == 1:
		button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
		button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

		button2 = tk.Button(frame_1, bg='gray', state='disabled')
		button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

		button3 = tk.Button(frame_1, bg='gray', state='disabled')
		button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

		button4 = tk.Button(frame_1, bg='gray', state='disabled')
		button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

		button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
		button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
	if num_of_jars == 0:
		button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
		button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

		button2 = tk.Button(frame_1, bg='gray', state='disabled')
		button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

		button3 = tk.Button(frame_1, bg='gray', state='disabled')
		button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

		button4 = tk.Button(frame_1, bg='gray', state='disabled')
		button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

		button5 = tk.Button(frame_1, bg='gray', state='disabled') # Rightmost button
		button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)


	#LEFT AND RIGHT BUTTONS

	if num_of_jars >= 6:
		button_left = tk.Button(frame_left, bg='#FF7000',text='<<', command=move_left)
		button_left.place(relwidth=1, relheight=1)
	else:
		button_left = tk.Button(frame_left, bg='gray',text='<<', command=move_left, state='disabled')
		button_left.place(relwidth=1, relheight=1)


	button_right = tk.Button(frame_right, bg='gray', text='>>', command=move_right, state ='disabled')
	button_right.place(relwidth=1, relheight=1)


	window.destroy()

def left_create_jar(root):
	global position_seed_provider

	position_seed_provider -= 1

	button_seed_provider = tk.Button(root, text=list_seed_providers[position_seed_provider], state='disabled')
	button_seed_provider.place(relx=0.56, rely=0.625+7*(0.0375), relwidth=0.3, relheight=0.075)

	if position_seed_provider == 0:
		button_left_provider= tk.Button(root, bg='gray', text='<<', state='disabled')
		button_left_provider.place(relx=0.45, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)

	button_right_provider = tk.Button(root, bg='orange', text='>>', state='normal',command=lambda: right_create_jar(root))
	button_right_provider.place(relx=0.87, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)

def right_create_jar(root):
	global position_seed_provider

	position_seed_provider += 1

	button_seed_provider = tk.Button(root, text=list_seed_providers[position_seed_provider], state='disabled')
	button_seed_provider.place(relx=0.56, rely=0.625+7*(0.0375), relwidth=0.3, relheight=0.075)

	if position_seed_provider == len(list_seed_providers)-1:
		button_right_provider= tk.Button(root, bg='gray', text='>>', state='disabled')
		button_right_provider.place(relx=0.87, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)

	button_left_provider = tk.Button(root, bg='orange', text='<<', state='normal',command=lambda: left_create_jar(root))
	button_left_provider.place(relx=0.45, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)

def add_provider(provider_name, cost, root_new_provider, root_create_jar):
	global list_seed_providers
	global position_seed_provider
	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)

	sheet = jar_database['Database']
	sheet2 = jar_database['Cost']

	ROW = 4
	COLUMN = 2

	cell = sheet2.cell(row=ROW, column=COLUMN).value

	while cell is not None:
		COLUMN += 1
		cell = sheet2.cell(row=ROW, column=COLUMN).value

	sheet2.cell(row=ROW, column=COLUMN, value=provider_name)
	sheet2.cell(row=ROW+1, column=COLUMN, value=float(cost))

	ROW = 4
	COLUMN = 2

	list_seed_providers = []
	print(list_seed_providers, "HEHFSDKJFHKSJDFSD")

	cell = sheet2.cell(row=ROW, column=COLUMN).value
	while cell is not None:
		list_seed_providers.append(cell)
		COLUMN += 1
		cell = sheet2.cell(row=ROW, column=COLUMN).value

	print(list_seed_providers, "HEHFSDKJFHKSJDFSD times 2")

	position_seed_provider = 0

	button_seed_provider = tk.Button(root_create_jar, text=list_seed_providers[position_seed_provider], state='disabled') # this is a button but doesnt actually work as a button. Prettu much a label
	button_seed_provider.place(relx=0.56, rely=0.625+7*(0.0375), relwidth=0.3, relheight=0.075)

	button_left_provider = tk.Button(root_create_jar, font=button_font, bg='gray', text='<<', state='disabled')
	button_left_provider.place(relx=0.45, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)

	if len(list_seed_providers) == 0 or len(list_seed_providers) == 1:
		button_right_provider = tk.Button(root_create_jar, font=button_font, bg='gray', text='>>', state='disabled')
		button_right_provider.place(relx=0.87, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)
	else:
		button_right_provider = tk.Button(root_create_jar, bg='orange', font=button_font, text='>>', command=lambda: right_create_jar(root_create_jar))
		button_right_provider.place(relx=0.87, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)



	jar_database.save(file_rel_path)

	root_new_provider.destroy()


def new_provider(root_create_jar):

	root_new_provider = tk.Tk()
	root_new_provider.title('New Provider')

	WIDTH = 450
	HEIGHT = 100

	label_font = 'Calibri 18 bold'

	canvas = tk.Canvas(root_new_provider, height=HEIGHT, width=WIDTH)
	canvas.pack()

	label_provider = tk.Label(root_new_provider, font=label_font, text='Provider:')
	label_provider.place(relx=0, rely=0, relwidth=1/4, relheight=1/2)

	label_cost = tk.Label(root_new_provider, font=label_font, text='Cost (tbsp):')
	label_cost.place(relx=0, rely=0.5, relwidth=1/4, relheight=1/2)


	entry_provider = tk.Entry(root_new_provider, font=label_font)
	entry_provider.place(relx=1/4, rely=0.1, relwidth=1/2, relheight=0.6/2)

	entry_cost = tk.Entry(root_new_provider, font=label_font)
	entry_cost.place(relx=1/4, rely=0.6, relwidth=1/2, relheight=0.6/2)


	button_create_provider = tk.Button(root_new_provider, font=label_font, text='Create', command= lambda: add_provider(entry_provider.get(), entry_cost.get(), root_new_provider, root_create_jar))
	button_create_provider.place(relx=25/32, rely=0.2, relwidth=6/32, relheight=0.6)

def create_jar(jar_num):
	global position_seed_provider
	global list_seed_providers

	root_create_jar = tk.Tk()
	root_create_jar.title('Create Jar')
	import time # Idk why this is here but im scared to remove it. It probably shouldnt be there. Ill double check later and remove it if it is not needed.
	new_jar_column = jar_num+2

	WIDTH = 350
	HEIGHT = 450

	canvas = tk.Canvas(root_create_jar, height=HEIGHT, width=WIDTH)
	canvas.pack()

	label_background=tk.Label(root_create_jar, bg='pink') # This label is empty it just contains the background color
	label_background.place(relwidth=1, relheight=1)

	entry_font = 'Calibri 18 bold'
	label_font = 'Calibri 12 bold'
	button_font = 'Calibri 10 bold'

	todays_date = datetime.date.today()
	todays_date = [str(todays_date.month), str(todays_date.day), str(todays_date.year)]
	todays_date = '/'.join(todays_date)


	label_date = tk.Label(root_create_jar, bg='pink', font=entry_font, text=todays_date)
	label_date.place(relx=0.2, rely=0.01, relwidth=0.6, relheight=0.08)


	# ALL ENTRIIES BELOW
	label_seed_amt = tk.Label(root_create_jar, font=label_font, text='Seed Amount:')
	label_seed_amt.place(relx=0.04, rely=0.1, relwidth=0.35, relheight=0.075)

	label_weight = tk.Label(root_create_jar, font=label_font, text='Weight:')
	label_weight.place(relx=0.04, rely=0.175+(0.0375), relwidth=0.35, relheight=0.075)

	label_floating_seeds = tk.Label(root_create_jar, font=label_font, text='Floating Seeds:')
	label_floating_seeds.place(relx=0.04, rely=0.25+2*(0.0375), relwidth= 0.35, relheight=0.075)

	label_daily_rinses = tk.Label(root_create_jar, font=label_font, text='Daily Rinses:')
	label_daily_rinses.place(relx=0.04, rely=0.325+3*(0.0375), relwidth= 0.35, relheight=0.075)

	label_dark_days = tk.Label(root_create_jar, font=label_font, text='Days In Dark:')
	label_dark_days.place(relx=0.04, rely=0.4+4*(0.0375), relwidth= 0.35, relheight=0.075)

	label_total_days = tk.Label(root_create_jar, font=label_font, text='Total Days:')
	label_total_days.place(relx=0.04, rely=0.475+5*(0.0375), relwidth= 0.35, relheight=0.075)

	label_initial_rinse = tk.Label(root_create_jar, font=label_font, text='Water (mL):')
	label_initial_rinse.place(relx=0.04, rely=0.55+6*(0.0375), relwidth= 0.35, relheight=0.075)

	label_seed_provider = tk.Label(root_create_jar, font=label_font, text='Seed Provider:')
	label_seed_provider.place(relx=0.04, rely=0.625+7*(0.0375), relwidth= 0.35, relheight=0.075)


	#ALL ENTRIES FOR FUNCTION BELOW
	entry_seed_amt = tk.Entry(root_create_jar,font=entry_font)
	entry_seed_amt.place(relx=0.45, rely=0.1, relwidth=0.3, relheight=0.075)

	entry_weight = tk.Entry(root_create_jar, font=entry_font)
	entry_weight.place(relx=0.45, rely=0.175+(0.0375), relwidth=0.3, relheight=0.075)

	entry_floating_seeds = tk.Entry(root_create_jar, font=entry_font)
	entry_floating_seeds.place(relx=0.45, rely=0.25+2*(0.0375), relwidth=0.3, relheight=0.075)

	entry_daily_rinses = tk.Entry(root_create_jar, font=entry_font)
	entry_daily_rinses.place(relx=0.45, rely=0.325+3*(0.0375), relwidth=0.3, relheight=0.075)

	entry_dark_days = tk.Entry(root_create_jar, font=entry_font)
	entry_dark_days.place(relx=0.45, rely=0.4+4*(0.0375), relwidth=0.3, relheight=0.075)

	entry_total_days = tk.Entry(root_create_jar, font=entry_font)
	entry_total_days.place(relx=0.45, rely=0.475+5*(0.0375), relwidth=0.3, relheight=0.075)

	entry_initial_rinse = tk.Entry(root_create_jar, font=entry_font)
	entry_initial_rinse.place(relx=0.45, rely=0.55+6*(0.0375), relwidth=0.3, relheight=0.075)


	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)

	sheet = jar_database['Database']
	sheet2 = jar_database['Cost']

	ROW = 4
	COLUMN = 2

	list_seed_providers = []
	print(list_seed_providers, "HEHFSDKJFHKSJDFSD")

	cell = sheet2.cell(row=ROW, column=COLUMN).value
	while cell is not None:
		list_seed_providers.append(cell)
		COLUMN += 1
		cell = sheet2.cell(row=ROW, column=COLUMN).value

	print(list_seed_providers, "HEHFSDKJFHKSJDFSD times 2")

	position_seed_provider = 0

	if len(list_seed_providers) == 0:
		button_seed_provider = tk.Button(root_create_jar, text='no providers', state='disabled') # this is a button but doesnt actually work as a button. Prettu much a label
		button_seed_provider.place(relx=0.56, rely=0.625+7*(0.0375), relwidth=0.3, relheight=0.075)
	else:
		button_seed_provider = tk.Button(root_create_jar, text=list_seed_providers[position_seed_provider], state='disabled') # this is a button but doesnt actually work as a button. Prettu much a label
		button_seed_provider.place(relx=0.56, rely=0.625+7*(0.0375), relwidth=0.3, relheight=0.075)

	#BUTTON
	button_create_provider = tk.Button(root_create_jar, text='New\nProvider', command=lambda: new_provider(root_create_jar))
	button_create_provider.place(relx=0.80, rely=0.55+6*(0.0375), relwidth=0.15, relheight=0.075)

	button_create = tk.Button(root_create_jar, font=button_font, text='CREATE', command=lambda: type_to_excel(todays_date, entry_seed_amt.get(), entry_weight.get(), entry_floating_seeds.get(), entry_daily_rinses.get(), entry_dark_days.get(), entry_total_days.get(), entry_initial_rinse.get(), new_jar_column, root_create_jar))
	button_create.place(relx=0.8, rely=0.325+3*(0.375/7), relwidth=0.15, relheight=0.075)

	button_left_provider = tk.Button(root_create_jar, font=button_font, bg='gray', text='<<', state='disabled')
	button_left_provider.place(relx=0.45, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)

	if len(list_seed_providers) == 0 or len(list_seed_providers) == 1:
		button_right_provider = tk.Button(root_create_jar, font=button_font, bg='gray', text='>>', state='disabled')
		button_right_provider.place(relx=0.87, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)
	else:
		button_right_provider = tk.Button(root_create_jar, bg='orange', font=button_font, text='>>', command=lambda: right_create_jar(root_create_jar))
		button_right_provider.place(relx=0.87, rely=0.625+7*(0.0375), relwidth=0.1, relheight=0.075)

def add_water(water, add_water_column, root, pos_plus_num): # called in the show_info function

	global jar_list

	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)
	sheet = jar_database['Database']

	current_water_value = float(sheet.cell(row=11, column=add_water_column).value)
	sheet.cell(row=11, column=add_water_column, value=current_water_value+float(water))
	new_water_value = sheet.cell(row=11, column=add_water_column).value



	# This will add a water entry to the excel sheet (record date, time, and amount of water added in the excel sheet.)
	now = datetime.datetime.now()

	date = now.strftime("%m/%d/%y")

	hour = now.strftime('%H')
	print(hour)
	minute = now.strftime(':%M')

	if int(hour) == 0:
		hour = '12'
		am_or_pm = ' AM'
	elif int(hour) == 12:
		am_or_pm = ' PM'
	elif int(hour) > 12:
		hour = str(int(hour) - 12)
		am_or_pm = ' PM'
	else:
		am_or_pm = ' AM'

	time = hour + minute
	time = time + am_or_pm
	date_and_time = date + ' - ' + time

	water_record = date_and_time + ' - ' + water + 'mL'


	water_entry_row = 21
	while bool(sheet.cell(row=water_entry_row, column=add_water_column).value) == True:
		water_entry_row += 1
	sheet.cell(row=water_entry_row, column=add_water_column, value=water_record)


	jar_database.save(file_rel_path) 

	jar_list[pos_plus_num][9] = new_water_value

	print(sheet.cell(row=11, column=add_water_column).value)

	label_font = 'Calibri 10 bold'
	label_water_used2 = tk.Label(root, font=label_font, text=jar_list[pos_plus_num][9], borderwidth=2, relief='ridge')
	label_water_used2.place(relx=0.45, rely=0.52+6*(0.41/9), relwidth= 0.35, relheight=0.07)

def final_weight(weight, final_weight_column, root, pos_plus_num):

	global jar_list

	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)
	sheet = jar_database['Database']
	print('POSITION', position)
	print('NUM OF JAR JARS: ', num_of_jars)
	print('DIS DA BUTTON NUM', pos_plus_num-position)





	sheet.cell(row=12, column=final_weight_column, value=float(weight)) # Adds the final weight to the excell sheet
	sheet.cell(row=3, column=final_weight_column, value='harvested') # Changes the status on the excel sheet only to harvested, not the jar_list. That is done below
	sheet.cell(row=4, column=final_weight_column, value='#FE7A74') # Changes the Entry Color on the excel sheet only to harvested, not the jar_list. That is done below
	print(sheet.cell(row=12, column=final_weight_column).value)
	print(sheet.cell(row=6, column=final_weight_column).value)
	growth = sheet.cell(row=12, column=final_weight_column).value / sheet.cell(row=6, column=final_weight_column).value
	print(growth)
	print(type(growth))
	sheet.cell(row=13, column=final_weight_column, value='%.2fx' % growth)

	sheet2 = jar_database['Cost']
	current_water_cost = sheet2.cell(row=2, column=2).value
	current_seed_cost = sheet.cell(row=15, column=final_weight_column).value ####PROBLEM IT IS LOOKING AT ONE PLACE WHEN IT SHOULD BE LOOKING DYNAMICALY DEPENDING ON PROVIDER

	# Writes down the cost of water and seed at time of harvest
	#### I CAN PROBABLY DELETE TWO LINES BELOW ####

	# Calculates price using the current prices of water and seeds and the amount of water and seeds used.
	print("URMOM12345")
	print(sheet.cell(row=11, column=final_weight_column).value)
	print(type(sheet.cell(row=11, column=final_weight_column).value))

	print(sheet.cell(row=14, column=final_weight_column).value)
	print(type(sheet.cell(row=14, column=final_weight_column).value))


	water_cost = (sheet.cell(row=11, column=final_weight_column).value / 1000) * sheet.cell(row=14, column=final_weight_column).value
	seed_cost = sheet.cell(row=5, column=final_weight_column).value * sheet.cell(row=15, column=final_weight_column).value

	total_cost = water_cost + seed_cost
	print(type(total_cost))
	# Writes down the prices into excel
	sheet.cell(row=16, column=final_weight_column, value=float('%.2f' % water_cost))
	sheet.cell(row=17, column=final_weight_column, value=float('%.2f' % seed_cost))
	sheet.cell(row=18, column=final_weight_column, value=float('%.2f' % total_cost))

	print("HSKUJDFHK")
	print(water_cost)
	print(type(water_cost))
	print(seed_cost)
	print(type(seed_cost))



	jar_database.save(file_rel_path)

	jar_list[pos_plus_num][1] = 'harvested' # Updates the status on the jar_list     #
	jar_list[pos_plus_num][2] = '#FE7A74' # Updates the entry color on the jar_list. ###This way program can recognize and change its color, status, and final weight without having to close and reopen the program
	jar_list[pos_plus_num][10] = weight #Updates the weight on the jar_list.         #
	jar_list[pos_plus_num][11] = str(growth)
	jar_list[pos_plus_num][12] = str(current_water_cost)
	jar_list[pos_plus_num][13] = str(current_seed_cost)
	jar_list[pos_plus_num][14] = str('%.2f' % water_cost)
	jar_list[pos_plus_num][15] = str('%.2f' % seed_cost)
	jar_list[pos_plus_num][16] = str('%.2f' % total_cost)

	print('GUIDO LOOK HERE: ', jar_list[pos_plus_num])
	print(jar_list[pos_plus_num][12])



	button_color_update = int(pos_plus_num - position)

	# Every time I update buttons, they use the list jar_list[position+NUM] where NUM is a value between 0 and 4. Each button (button1, button2) always has the same value. (button1 is always position+0, button2 is always position+1) So if i get pos_plus_num - position , that will give me that NUM thats unique to each button.(this is cause pos_plus_num is equal to the position plus that number (position+NUM)) so pos_plus_num - position is just like saying      position + NUM - position. Positions cancel out leaving only that NUM. That num will tell me which button was clicked. Sorry for this longass explination

	if button_color_update ==  4:
		button1 = tk.Button(frame_1, bg=jar_list[position+4][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+4][0],jar_list[position+4][1],(datetime.date.today()-jar_dates[position+4]).days,jar_list[position+4][8]), command=lambda: show_info(jar_list[position+4], position+4)) # Leftmost button
		button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)
	if button_color_update == 3:
		button2 = tk.Button(frame_1, bg=jar_list[position+3][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+3][0],jar_list[position+3][1],(datetime.date.today()-jar_dates[position+3]).days,jar_list[position+3][8]), command=lambda: show_info(jar_list[position+3], position+3))
		button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)
	if button_color_update == 2:
		button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
		button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)
	if button_color_update == 1:
		button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
		button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)
	if button_color_update == 0:
		button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
		button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)

	print(sheet.cell(row=12, column=final_weight_column).value)

	root.destroy()

def show_info(jar_list_index, pos_plus_num):
	root_show_info = tk.Tk()

	column_show_info = num_of_jars+1-pos_plus_num


	

	if jar_list_index[1] == 'growing':
		root_show_info.title('Growing!')

		WIDTH = 350
		HEIGHT = 500

		canvas = tk.Canvas(root_show_info, height=HEIGHT, width=WIDTH)
		canvas.pack()

		label_background = tk.Label(root_show_info, bg='#85DE77')
		label_background.place(relwidth=1, relheight=1)

		date_font = 'Calibri 16 bold'
		label_font = 'Calibri 10 bold'
		button_font = 'Calibri 10 bold'

		date_of_jar = jar_list_index[0]

		label_date = tk.Label(root_show_info, bg='#85DE77', font=date_font, text=date_of_jar)
		label_date.place(relx=0.2, rely=0.01, relwidth=0.6, relheight=0.08)


		# ALL ENTRIIES BELOW
		label_seed_amt = tk.Label(root_show_info, font=label_font, text='Seed Amount (tbsp):', borderwidth=2, relief='ridge')
		label_seed_amt.place(relx=0.04, rely=0.1, relwidth=0.35, relheight=0.07)

		label_weight = tk.Label(root_show_info, font=label_font, text='Weight (g):', borderwidth=2, relief='ridge')
		label_weight.place(relx=0.04, rely=0.17+(0.41/9), relwidth=0.35, relheight=0.07)

		label_floating_seeds = tk.Label(root_show_info, font=label_font, text='Floating Seeds:', borderwidth=2, relief='ridge')
		label_floating_seeds.place(relx=0.04, rely=0.24+2*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_daily_rinses = tk.Label(root_show_info, font=label_font, text='Daily Rinses:', borderwidth=2, relief='ridge')
		label_daily_rinses.place(relx=0.04, rely=0.31+3*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_dark_days = tk.Label(root_show_info, font=label_font, text='Days In Dark:', borderwidth=2, relief='ridge')
		label_dark_days.place(relx=0.04, rely=0.38+4*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_total_days = tk.Label(root_show_info, font=label_font, text='Total Days:', borderwidth=2, relief='ridge')
		label_total_days.place(relx=0.04, rely=0.45+5*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_water_used = tk.Label(root_show_info, font=label_font, text='Water (mL):', borderwidth=2, relief='ridge')
		label_water_used.place(relx=0.04, rely=0.52+6*(0.41/9), relwidth= 0.35, relheight=0.07)




		#ALL ENTRIES FOR FUNCTION BELOW
		label_seed_amt2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[3], borderwidth=2, relief='ridge')
		label_seed_amt2.place(relx=0.45, rely=0.1, relwidth=0.35, relheight=0.07)

		label_weight2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[4], borderwidth=2, relief='ridge')
		label_weight2.place(relx=0.45, rely=0.17+(0.41/9), relwidth=0.35, relheight=0.07)

		label_floating_seeds2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[5], borderwidth=2, relief='ridge')
		label_floating_seeds2.place(relx=0.45, rely=0.24+2*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_daily_rinses2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[6], borderwidth=2, relief='ridge')
		label_daily_rinses2.place(relx=0.45, rely=0.31+3*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_dark_days2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[7], borderwidth=2, relief='ridge')
		label_dark_days2.place(relx=0.45, rely=0.38+4*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_total_days2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[8], borderwidth=2, relief='ridge')
		label_total_days2.place(relx=0.45, rely=0.45+5*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_water_used2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[9], borderwidth=2, relief='ridge')
		label_water_used2.place(relx=0.45, rely=0.52+6*(0.41/9), relwidth= 0.35, relheight=0.07)


		# Label, Entry, and Button to add water

		label_add_water = tk.Label(root_show_info, bg='#85DE77', font='Calibri 12 bold', text='+')
		label_add_water.place(relx=0.04, rely=0.59+6*(0.41/9), relwidth=0.35, relheight=0.05)

		entry_add_water = tk.Entry(root_show_info, font=label_font)
		entry_add_water.place(relx=0.04, rely=0.59+7*(0.41/9), relwidth=0.35, relheight=0.07)

		button_add_water = tk.Button(root_show_info, font=button_font, text='Add Water (mL)', command=lambda: add_water(entry_add_water.get(), column_show_info, root_show_info, pos_plus_num))
		button_add_water.place(relx=0.45, rely=0.59+7*(0.41/9), relwidth=0.35, relheight=0.07)

	elif jar_list_index[1] == 'ready':
		root_show_info.title('Ready!')

		WIDTH = 350
		HEIGHT = 500

		canvas = tk.Canvas(root_show_info, height=HEIGHT, width=WIDTH)
		canvas.pack()

		label_background = tk.Label(root_show_info, bg='#FDFD98')
		label_background.place(relwidth=1, relheight=1)

		date_font = 'Calibri 16 bold'
		label_font = 'Calibri 10 bold'

		date_of_jar = jar_list_index[0]

		label_date = tk.Label(root_show_info, bg='#FDFD98', font=date_font, text=date_of_jar)
		label_date.place(relx=0.2, rely=0.01, relwidth=0.6, relheight=0.08)


		# ALL ENTRIIES BELOW
		label_seed_amt = tk.Label(root_show_info, font=label_font, text='Seed Amount (tbsp):', borderwidth=2, relief='ridge')
		label_seed_amt.place(relx=0.04, rely=0.1, relwidth=0.35, relheight=0.07)

		label_weight = tk.Label(root_show_info, font=label_font, text='Weight (g):', borderwidth=2, relief='ridge')
		label_weight.place(relx=0.04, rely=0.17+(0.41/9), relwidth=0.35, relheight=0.07)

		label_floating_seeds = tk.Label(root_show_info, font=label_font, text='Floating Seeds:', borderwidth=2, relief='ridge')
		label_floating_seeds.place(relx=0.04, rely=0.24+2*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_daily_rinses = tk.Label(root_show_info, font=label_font, text='Daily Rinses:', borderwidth=2, relief='ridge')
		label_daily_rinses.place(relx=0.04, rely=0.31+3*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_dark_days = tk.Label(root_show_info, font=label_font, text='Days In Dark:', borderwidth=2, relief='ridge')
		label_dark_days.place(relx=0.04, rely=0.38+4*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_total_days = tk.Label(root_show_info, font=label_font, text='Total Days:', borderwidth=2, relief='ridge')
		label_total_days.place(relx=0.04, rely=0.45+5*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_water_used = tk.Label(root_show_info, font=label_font, text='Water (mL):', borderwidth=2, relief='ridge')
		label_water_used.place(relx=0.04, rely=0.52+6*(0.41/9), relwidth= 0.35, relheight=0.07)




		#ALL ENTRIES FOR FUNCTION BELOW
		label_seed_amt2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[3], borderwidth=2, relief='ridge')
		label_seed_amt2.place(relx=0.45, rely=0.1, relwidth=0.35, relheight=0.07)

		label_weight2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[4], borderwidth=2, relief='ridge')
		label_weight2.place(relx=0.45, rely=0.17+(0.41/9), relwidth=0.35, relheight=0.07)

		label_floating_seeds2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[5], borderwidth=2, relief='ridge')
		label_floating_seeds2.place(relx=0.45, rely=0.24+2*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_daily_rinses2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[6], borderwidth=2, relief='ridge')
		label_daily_rinses2.place(relx=0.45, rely=0.31+3*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_dark_days2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[7], borderwidth=2, relief='ridge')
		label_dark_days2.place(relx=0.45, rely=0.38+4*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_total_days2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[8], borderwidth=2, relief='ridge')
		label_total_days2.place(relx=0.45, rely=0.45+5*(0.41/9), relwidth= 0.35, relheight=0.07)

		label_water_used2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[9], borderwidth=2, relief='ridge')
		label_water_used2.place(relx=0.45, rely=0.52+6*(0.41/9), relwidth= 0.35, relheight=0.07)


		# Label, Entry, and Button to add final weight

		label_final_weight = tk.Label(root_show_info, bg='#FDFD98', font=label_font, text='final weight (g)')
		label_final_weight.place(relx=0.04, rely=0.60+6*(0.41/9), relwidth=0.35, relheight=0.05)

		entry_final_weight = tk.Entry(root_show_info, font=label_font)
		entry_final_weight.place(relx=0.04, rely=0.59+7*(0.41/9), relwidth=0.35, relheight=0.07)

		button_final_weight = tk.Button(root_show_info, font=label_font, text='Harvest!', command=lambda: final_weight(entry_final_weight.get(), column_show_info, root_show_info, pos_plus_num))
		button_final_weight.place(relx=0.45, rely=0.59+7*(0.41/9), relwidth=0.35, relheight=0.07)

	elif jar_list_index[1] == 'harvested':
		root_show_info.title('Harvested!')

		WIDTH = 850
		HEIGHT = 600

		canvas = tk.Canvas(root_show_info, height=HEIGHT, width=WIDTH)
		canvas.pack()

		label_background = tk.Label(root_show_info, bg='#FE7A74')
		label_background.place(relwidth=1, relheight=1)

		date_font = 'Calibri 20 bold'
		label_font = 'Calibri 14 bold'

		date_of_jar = jar_list_index[0]

		# Title Labels
		label_date = tk.Label(root_show_info, bg='#FE7A74', font=date_font, text=date_of_jar)
		label_date.place(relx=0.0, rely=0.01, relwidth=0.5, relheight=0.08)

		label_cost = tk.Label(root_show_info, bg='#FE7A74', font=date_font, text='COST')
		label_cost.place(relx=0.5, rely=0.01, relwidth=0.5, relheight=0.08)

		label_picture = tk.Label(root_show_info, bg='#FE7A74', font=date_font, text='PICTURE')
		label_picture.place(relx=0.5, rely=0.1+5*(0.9/13)+5*(3.4/130), relwidth=0.5, relheight=0.08)		


		# All on the left side of left screen
		label_seed_amt = tk.Label(root_show_info, font=label_font, text='Seed Amount (tbsp):', borderwidth=2, relief='ridge')
		label_seed_amt.place(relx=0.03, rely=0.1, relwidth=0.20, relheight=0.9/13)

		label_weight = tk.Label(root_show_info, font=label_font, text='Weight (g):', borderwidth=2, relief='ridge')
		label_weight.place(relx=0.03, rely=0.1+(0.9/13)+(3.6/130), relwidth=0.20, relheight=0.9/13)

		label_floating_seeds = tk.Label(root_show_info, font=label_font, text='Floating Seeds:', borderwidth=2, relief='ridge')
		label_floating_seeds.place(relx=0.03, rely=0.1+2*(0.9/13)+2*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_daily_rinses = tk.Label(root_show_info, font=label_font, text='Daily Rinses:', borderwidth=2, relief='ridge')
		label_daily_rinses.place(relx=0.03, rely=0.1+3*(0.9/13)+3*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_dark_days = tk.Label(root_show_info, font=label_font, text='Days In Dark:', borderwidth=2, relief='ridge')
		label_dark_days.place(relx=0.03, rely=0.1+4*(0.9/13)+4*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_total_days = tk.Label(root_show_info, font=label_font, text='Total Days:', borderwidth=2, relief='ridge')
		label_total_days.place(relx=0.03, rely=0.1+5*(0.9/13)+5*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_water_used = tk.Label(root_show_info, font=label_font, text='Water (mL):', borderwidth=2, relief='ridge')
		label_water_used.place(relx=0.03, rely=0.1+6*(0.9/13)+6*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_final_weight = tk.Label(root_show_info, font=label_font, text='Final Weight (g):', borderwidth=2, relief='ridge')
		label_final_weight.place(relx=0.03, rely=0.1+7*(0.9/13)+7*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_weight_growth = tk.Label(root_show_info, font=label_font, text='Weight Growth:', borderwidth=2, relief='ridge')
		label_weight_growth.place(relx=0.03, rely=0.1+8*(0.9/13)+8*(3.6/130), relwidth= 0.20, relheight=0.9/13)


		# All on the right side of left screen
		label_seed_amt2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[3], borderwidth=2, relief='ridge')
		label_seed_amt2.place(relx=0.27, rely=0.1, relwidth=0.20, relheight=0.9/13)

		label_weight2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[4], borderwidth=2, relief='ridge')
		label_weight2.place(relx=0.27, rely=0.1+(0.9/13)+(2.7/130), relwidth=0.20, relheight=0.9/13)

		label_floating_seeds2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[5], borderwidth=2, relief='ridge')
		label_floating_seeds2.place(relx=0.27, rely=0.1+2*(0.9/13)+2*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_daily_rinses2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[6], borderwidth=2, relief='ridge')
		label_daily_rinses2.place(relx=0.27, rely=0.1+3*(0.9/13)+3*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_dark_days2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[7], borderwidth=2, relief='ridge')
		label_dark_days2.place(relx=0.27, rely=0.1+4*(0.9/13)+4*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_total_days2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[8], borderwidth=2, relief='ridge')
		label_total_days2.place(relx=0.27, rely=0.1+5*(0.9/13)+5*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_water_used2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[9], borderwidth=2, relief='ridge')
		label_water_used2.place(relx=0.27, rely=0.1+6*(0.9/13)+6*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_final_weight2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[10], borderwidth=2, relief='ridge')
		label_final_weight2.place(relx=0.27, rely=0.1+7*(0.9/13)+7*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		label_weight_growth2 = tk.Label(root_show_info, font=label_font, text=jar_list_index[11], borderwidth=2, relief='ridge')
		label_weight_growth2.place(relx=0.27, rely=0.1+8*(0.9/13)+8*(3.6/130), relwidth= 0.20, relheight=0.9/13)

		# Labels for costs
		label_water_cost_per_L = tk.Label(root_show_info, bg='#FE7A74', font=label_font, text='Water Cost Per L:', borderwidth=2, relief='ridge')
		label_water_cost_per_L.place(relx=0.53, rely=0.1, relwidth=0.20, relheight=0.9/13)

		label_seed_cost_per_tbsp = tk.Label(root_show_info, bg='#FE7A74', font=label_font, text='Seed Cost Per tbsp:', borderwidth=2, relief='ridge')
		label_seed_cost_per_tbsp.place(relx=0.53, rely=0.1+(0.9/13)+(2.7/130), relwidth=0.20, relheight=0.9/13)

		label_water_cost_per_L2 = tk.Label(root_show_info, bg='#FE7A74', font=label_font, text='%s' % '$' + str(jar_list_index[12]), borderwidth=2, relief='ridge')
		label_water_cost_per_L2.place(relx=0.77, rely=0.1, relwidth=0.20, relheight=0.9/13)

		label_seed_cost_per_tbsp2 = tk.Label(root_show_info, bg='#FE7A74', font=label_font, text='%s' % '$' + str(jar_list_index[13]), borderwidth=2, relief='ridge')
		label_seed_cost_per_tbsp2.place(relx=0.77, rely=0.1+(0.9/13)+(2.7/130), relwidth=0.20, relheight=0.9/13)


		label_water_cost = tk.Label(root_show_info, font=label_font, text='Jar Water Cost:', borderwidth=2, relief='ridge')
		label_water_cost.place(relx=0.53, rely=0.1+2*(0.9/13)+2*(3.6/130), relwidth=0.20, relheight=0.9/13)

		label_seed_cost = tk.Label(root_show_info, font=label_font, text='Jar Seed Cost:', borderwidth=2, relief='ridge')
		label_seed_cost.place(relx=0.53, rely=0.1+3*(0.9/13)+3*(3.6/130), relwidth=0.20, relheight=0.9/13)

		label_water_cost2 = tk.Label(root_show_info, font=label_font, text='%s' % '$' + str(jar_list_index[14]), borderwidth=2, relief='ridge')
		label_water_cost2.place(relx=0.77, rely=0.1+2*(0.9/13)+2*(3.6/130), relwidth=0.20, relheight=0.9/13)

		label_seed_cost2 = tk.Label(root_show_info, font=label_font, text='%s' % '$' + str(jar_list_index[15]), borderwidth=2, relief='ridge')
		label_seed_cost2.place(relx=0.77, rely=0.1+3*(0.9/13)+3*(3.6/130), relwidth=0.20, relheight=0.9/13)


		label_total_cost = tk.Label(root_show_info, font=label_font, text='Total Cost:', borderwidth=2, relief='ridge')
		label_total_cost.place(relx=0.53, rely=0.1+4*(0.9/13)+4*(3.6/130), relwidth=0.20, relheight=0.9/13)

		label_total_cost2 = tk.Label(root_show_info, font=label_font, text='%s' % '$' + str(jar_list_index[16]), borderwidth=2, relief='ridge')
		label_total_cost2.place(relx=0.77, rely=0.1+4*(0.9/13)+4*(3.6/130), relwidth=0.20, relheight=0.9/13)

		# Picture Button
		button_picture = tk.Button(root_show_info, font=label_font, text='Click to see\nPicture', borderwidth=8)
		button_picture.place(relx=0.53, rely=0.1+6*(0.9/13)+6*(3.6/130), relwidth=0.44, relheight=3*(0.9/13)+2*(3.6/130))

def exit_full_screen(event):
	root.attributes('-fullscreen', False)

def water_button(price, root, entry):
	global screen_water_cost

	# Deletes the entry box content
	entry.delete(0, 'end')
	label_font = 'Calibri 16 bold'

	screen_water_cost = price

	# Updates the label with the new price
	label = tk.Label(root, bg='sky blue', font=label_font, text='Water: $%.2f' % price)
	label.place(relx=0.3, rely=0.05, relwidth=0.4, relheight=0.05)

	# Writes new price on
	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)

	sheet2 = jar_database['Cost']
	sheet2.cell(row=2, column=2, value=price)

	button_price = tk.Button(frame_price, bg='#FE7A74', font='Calibri 14 bold', text='Water: $%.2f/L\n\n%s' % (screen_water_cost, 'Change Costs'), command=lambda: change_price(screen_water_cost))
	button_price.place(relwidth=1, relheight=1)

	jar_database.save(file_rel_path) 

def seed_button(price, position, root, entry):

	# Deletes the entry box content
	entry.delete(0, 'end')
	label_font = 'Calibri 16 bold'

	#screen_seed_cost = price

	# Updates the label with the new price
	label = tk.Label(root, bg='sky blue', font=label_font, text='Price changed to $%.2f' % price)
	label.place(relx=0.65, rely=0.56, relwidth=0.3, relheight=0.05)

	# Writes new price on
	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)

	sheet2 = jar_database['Cost']
	sheet2.cell(row=5, column=position+2, value=price)

	#button_price = tk.Button(frame_price, bg='#FE7A74', font='Calibri 14 bold', text='Water: $%.2f/L\n\n%s' % (screen_water_cost, 'Change Costs'), command=lambda: change_price(screen_water_cost))
	#utton_price.place(relwidth=1, relheight=1)  ### NOT NEEDED SINCE SEED PRICE CHANGING DOES NOT CHANGE HOW THE BUTTON LOOKS AS SEED PRICE IS NOT SHOWN IN BUTTON

	jar_database.save(file_rel_path) 

def right_change_price(root):
	global position_seed_provider
	global list_seed_providers

	position_seed_provider += 1

	button_seed_provider = tk.Button(root, text=list_seed_providers[position_seed_provider], state='disabled')
	button_seed_provider.place(relx=0.3, rely=0.46, relwidth=0.4, relheight=0.09)

	if position_seed_provider == len(list_seed_providers)-1:
		button_right_provider= tk.Button(root, bg='gray', text='>>', state='disabled')
		button_right_provider.place(relx=0.72, rely=0.46, relwidth=0.08, relheight=0.09)

	button_left_provider = tk.Button(root, bg='orange', text='<<', state='normal',command=lambda: left_change_price(root))
	button_left_provider.place(relx=0.2, rely=0.46, relwidth=0.08, relheight=0.09)
	print(position_seed_provider)

def left_change_price(root):
	global position_seed_provider
	global list_seed_providers

	position_seed_provider -= 1

	button_seed_provider = tk.Button(root, text=list_seed_providers[position_seed_provider], state='disabled')
	button_seed_provider.place(relx=0.3, rely=0.46, relwidth=0.4, relheight=0.09)

	if position_seed_provider == 0:
		button_left_provider= tk.Button(root, bg='gray', text='<<', state='disabled')
		button_left_provider.place(relx=0.2, rely=0.46, relwidth=0.08, relheight=0.09)

	button_right_provider = tk.Button(root, bg='orange', text='>>', state='normal',command=lambda: right_change_price(root))
	button_right_provider.place(relx=0.72, rely=0.46, relwidth=0.08, relheight=0.09)
	print(position_seed_provider)

def change_price(screen_water_cost):
	global position_seed_provider
	global list_seed_providers

	root_change_price = tk.Tk()
	root_change_price.title('Prices')
	WIDTH = 400
	HEIGHT = 400

	label_font = 'Calibri 16 bold'

	canvas = tk.Canvas(root_change_price, width=WIDTH, height=HEIGHT)
	canvas.pack()

	label_background = tk.Label(root_change_price, bg='sky blue')
	label_background.place(relwidth=1, relheight=1)

	label_water = tk.Label(root_change_price, bg='sky blue', font=label_font, text='Water: $%.2f' % screen_water_cost)
	label_water.place(relx=0.3, rely=0.05, relwidth=0.4, relheight=0.05)

	label_dollar_water = tk.Label(root_change_price, bg='sky blue', font=label_font, text='$')
	label_dollar_water.place(relx=0.35, rely=0.11, relwidth=0.05, relheight=0.10)

	#label_seed = tk.Label(root_change_price, bg='sky blue', font=label_font, text='Seeds: $%.2f' % screen_seed_cost) # THIS WILL BE REPLACED BY PROVIDER NAMES
	#label_seed.place(relx=0.3, rely=0.5, relwidth=0.4, relheight=0.05

	jar_database = openpyxl.load_workbook(file_rel_path, data_only=True)

	sheet = jar_database['Database']
	sheet2 = jar_database['Cost']

	ROW = 4
	COLUMN = 2

	list_seed_providers = []
	print(list_seed_providers, "HEHFSDKJFHKSJDFSD")

	cell = sheet2.cell(row=ROW, column=COLUMN).value
	while cell is not None:
		list_seed_providers.append(cell)
		COLUMN += 1
		cell = sheet2.cell(row=ROW, column=COLUMN).value

	print(list_seed_providers, "HEHFSDKJFHKSJDFSD times 2")

	position_seed_provider = 0

	if len(list_seed_providers) == 0:
		button_seed_provider = tk.Button(root_change_price, text='no providers', state='disabled') # this is a button but doesnt actually work as a button. Prettu much a label
		button_seed_provider.place(relx=0.3, rely=0.46, relwidth=0.4, relheight=0.09)
	else:
		button_seed_provider = tk.Button(root_change_price, text=list_seed_providers[position_seed_provider], state='disabled') # this is a button but doesnt actually work as a button. Prettu much a label
		button_seed_provider.place(relx=0.3, rely=0.46, relwidth=0.4, relheight=0.09)

	##############

	button_left_provider = tk.Button(root_change_price, font=button_font, bg='gray', text='<<', state='disabled')
	button_left_provider.place(relx=0.2, rely=0.46, relwidth=0.08, relheight=0.09)

	if len(list_seed_providers) == 0 or len(list_seed_providers) == 1:
		button_right_provider = tk.Button(root_change_price, font=button_font, bg='gray', text='>>', state='disabled')
		button_right_provider.place(relx=0.72, rely=0.46, relwidth=0.08, relheight=0.09)
	else:
		button_right_provider = tk.Button(root_change_price, bg='orange', font=button_font, text='>>', command=lambda: right_change_price(root_change_price))
		button_right_provider.place(relx=0.72, rely=0.46, relwidth=0.08, relheight=0.09)

#####
#####
#####

	label_dollar_seed = tk.Label(root_change_price, bg='sky blue', font=label_font, text='$')
	label_dollar_seed.place(relx=0.35, rely=0.56, relwidth=0.05, relheight=0.10)

	entry_water = tk.Entry(root_change_price, font=label_font)
	entry_water.place(relx=0.4, rely=0.11, relwidth=0.2, relheight=0.10)

	entry_seed = tk.Entry(root_change_price, font=label_font)
	entry_seed.place(relx=0.4, rely=0.56, relwidth=0.2, relheight=0.10)


	button_change_water = tk.Button(root_change_price, font=label_font, text='CHANGE WATER PRICE', command=lambda: water_button(float(entry_water.get()), root_change_price, entry_water))
	button_change_water.place(relx=0.24, rely=0.25, relwidth=0.52, relheight=0.20)

	button_change_seed = tk.Button(root_change_price, font=label_font, text='CHANGE SEED PRICE', command=lambda: seed_button(float(entry_seed.get()), position_seed_provider, root_change_price, entry_seed))
	button_change_seed.place(relx=0.24, rely=0.70, relwidth=0.52, relheight=0.20)

def show_all():
	root_show_all = tk.Tk()
	root_show_all.title('All Jars')

	WIDTH = 1200
	HEIGHT = 800

	label_font = 'Calibri 16 bold'

	canvas = tk.Canvas(root_show_all, width=WIDTH, height=HEIGHT, bg='sky blue')
	canvas.pack()

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.01, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.11, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.21, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.31, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.41, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.51, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.61, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.71, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.81, rely = 0.01, relwidth=0.08, relheight=0.04)

	button_test1= tk.Button(root_show_all, bg='yellow')
	button_test1.place(relx=0.91, rely = 0.01, relwidth=0.08, relheight=0.04)

def left_catagory(username):
	global position_catagory


	position_catagory -= 1

	button_catagory = tk.Button(root_catagory, text=catagory_list[position_catagory], command=lambda: open_catagory(catagory_list[position_catagory], username))
	button_catagory.place(relx=0.35, rely=0.35, relwidth=0.30, relheight=0.30)

	if position_catagory == 0:
		button_left_catagory = tk.Button(root_catagory, bg='gray', text='<<', state='disabled')
		button_left_catagory.place(relx=0.2, rely=0.425, relwidth=0.10, relheight=0.15)

	button_right_catagory = tk.Button(root_catagory, bg='orange', text='>>', state='normal', command=lambda: right_catagory(username))
	button_right_catagory.place(relx=0.7, rely=0.425, relwidth=0.10, relheight=0.15)

def right_catagory(username):
	global position_catagory

	position_catagory += 1

	button_catagory = tk.Button(root_catagory, text=catagory_list[position_catagory], command=lambda: open_catagory(catagory_list[position_catagory], username))
	button_catagory.place(relx=0.35, rely=0.35, relwidth=0.30, relheight=0.30)

	if position_catagory == len(catagory_list)-1:
		button_right_catagory = tk.Button(root_catagory, bg='gray', text='>>', state='disabled')
		button_right_catagory.place(relx=0.7, rely=0.425, relwidth=0.10, relheight=0.15)

	button_left_catagory = tk.Button(root_catagory, bg='orange', text='<<', state='normal',command=lambda: left_catagory(username))
	button_left_catagory.place(relx=0.2, rely=0.425, relwidth=0.10, relheight=0.15)

def add_catagory_to_dir(catagory_name, root_catagory, root_create_catagory, username):
	global catagory_list
	global position_catagory

	root_create_catagory.destroy()

	new_catagory_dir = os.path.join('users', username, catagory_name)
	new_catagory = os.mkdir(new_catagory_dir)

	catagory_dir = os.path.join('users', username) # This gets the new list of dirs in a users file
	catagory_list = os.listdir(catagory_dir)

	position_catagory = 0

	button_catagory = tk.Button(root_catagory, text=catagory_list[position_catagory], command=lambda: open_catagory(catagory_list[position_catagory], username))
	button_catagory.place(relx=0.35, rely=0.35, relwidth=0.30, relheight=0.30)

	if len(catagory_list) == 1:
		button_left_catagory = tk.Button(root_catagory, bg='gray', text='<<', state='disabled')
		button_left_catagory.place(relx=0.2, rely=0.425, relwidth=0.10, relheight=0.15)

		button_right_catagory = tk.Button(root_catagory, bg='gray', text='>>', state='disabled')
		button_right_catagory.place(relx=0.7, rely=0.425, relwidth=0.10, relheight=0.15)
	else:
		button_left_catagory = tk.Button(root_catagory, bg='gray', text='<<', state='disabled')
		button_left_catagory.place(relx=0.2, rely=0.425, relwidth=0.10, relheight=0.15)

		button_right_catagory = tk.Button(root_catagory, bg='orange', text='>>', command=lambda: right_catagory(username))
		button_right_catagory.place(relx=0.7, rely=0.425, relwidth=0.10, relheight=0.15)


def create_catagory(root_catagory, username):
	root_create_catagory = tk.Tk()
	root_create_catagory.title('Create Catagory')
	root_create_catagory.iconbitmap('broccoli.ico')

	canvas_create_catagory = tk.Canvas(root_create_catagory, width=300, height=150)
	canvas_create_catagory.pack()

	entry_create_catagory = tk.Entry(root_create_catagory)
	entry_create_catagory.place(relx=0.2, rely=0.20, relwidth=0.6, relheight=0.2)

	button_create_catagory = tk.Button(root_create_catagory, text='CREATE', command=lambda: add_catagory_to_dir(entry_create_catagory.get(), root_catagory, root_create_catagory, username))
	button_create_catagory.place(relx=0.2, rely=0.50, relwidth=0.6, relheight=0.2)

def login(username): # SWITCH UP TO A BUTTON-BASED CHOOSING THING SO ITS BUTTONS THAT HAVE THE NAME OF EACH FILE, NOT DROP-DOWN MENUES
	global position_catagory
	global root_catagory
	global catagory_list

	root_acc.destroy()
	root_catagory = tk.Tk()
	root_catagory.title(username)
	root_catagory.iconbitmap('broccoli.ico')

	catagory_dir = os.path.join('users', username)
	catagory_list = os.listdir(catagory_dir)

	canvas = tk.Canvas(root_catagory, bg='blue', width=400, height=250)
	canvas.pack()

	button_create_catagory = tk.Button(root_catagory, text='Crete new catagory', command=lambda: create_catagory(root_catagory, username))
	button_create_catagory.place(relx=0.1, rely=0.1, relwidth=0.8, relheight=0.20)

	position_catagory = 0

	if bool(catagory_list) == True:
		button_catagory = tk.Button(root_catagory, text=catagory_list[position_catagory], command=lambda: open_catagory(catagory_list[position_catagory], username))
		button_catagory.place(relx=0.35, rely=0.35, relwidth=0.30, relheight=0.30)

		if len(catagory_list) == 1:
			button_left_catagory = tk.Button(root_catagory, bg='gray', text='<<', state='disabled')
			button_left_catagory.place(relx=0.2, rely=0.425, relwidth=0.10, relheight=0.15)

			button_right_catagory = tk.Button(root_catagory, bg='gray', text='>>', state='disabled')
			button_right_catagory.place(relx=0.7, rely=0.425, relwidth=0.10, relheight=0.15)
		else:
			button_left_catagory = tk.Button(root_catagory, bg='gray', text='<<', state='disabled')
			button_left_catagory.place(relx=0.2, rely=0.425, relwidth=0.10, relheight=0.15)

			button_right_catagory = tk.Button(root_catagory, bg='orange', text='>>', command=lambda: right_catagory(username))
			button_right_catagory.place(relx=0.7, rely=0.425, relwidth=0.10, relheight=0.15)

	else:
		button_catagory = tk.Button(root_catagory, bg='gray', text='no catagory found', state='disabled')
		button_catagory.place(relx=0.35, rely=0.35, relwidth=0.30, relheight=0.30)

		button_left_catagory = tk.Button(root_catagory, bg='gray', text='<<', state='disabled')
		button_left_catagory.place(relx=0.2, rely=0.425, relwidth=0.10, relheight=0.15)

		button_right_catagory = tk.Button(root_catagory, bg='gray', text='>>', state='disabled')
		button_right_catagory.place(relx=0.7, rely=0.425, relwidth=0.10, relheight=0.15)




	print(username)
	user_dir = os.path.join('users', username)
	catagories = os.listdir(user_dir)
	print(catagories)
	print(bool(catagories))

def add_user_to_dir(username, root): # Later this will take in a root, username, and a password
	file_name = os.path.join('users', username) #this creates the file relative path/name that will be used to actually create the folder
	os.mkdir(file_name)
	print(file_name)
	user_list = os.listdir('users')
	account_ammount = len(user_list)
	print(account_ammount)
	if account_ammount == 1:
		button_acc_1 = tk.Button(frame_acc_names, text=username, command=lambda: login(username))
		button_acc_1.place(relwidth=1, relheight=0.2)
	elif account_ammount == 2:
		button_acc_2 = tk.Button(frame_acc_names, text=username, command=lambda: login(username))
		button_acc_2.place(rely=0.2, relwidth=1, relheight=0.2)
	elif account_ammount == 3:
		button_acc_3 = tk.Button(frame_acc_names, text=username, command=lambda: login(username))
		button_acc_3.place(rely=0.4, relwidth=1, relheight=0.2)
	elif account_ammount == 4:
		button_acc_4 = tk.Button(frame_acc_names, text=username, command=lambda: login(username))
		button_acc_4.place(rely=0.6, relwidth=1, relheight=0.2)
	elif account_ammount == 5:
		button_acc_5 = tk.Button(frame_acc_names, text=username, command=lambda: login(username))
		button_acc_5.place(rely=0.8, relwidth=1, relheight=0.2)

	if account_ammount == 5:
		button_create_account = tk.Button(frame_create_acc, text='CREATE\nACCOUNT', state= 'disabled', command= create_account)
		button_create_account.place(relwidth=1, relheight=1)

	root.destroy()

def create_account():
	root_create_account = tk.Tk()
	root_create_account.title('Create Sprouter Account')
	root_create_account.iconbitmap('broccoli.ico')

	font = 'Calibri 18'
	canvas_create_account = tk.Canvas(root_create_account,width=450, height=150)
	canvas_create_account.pack()

	# Labels
	label_username = tk.Label(root_create_account, font=font, text='Username:')
	label_username.place(rely=0.24, relwidth=0.25, relheight=0.25)

	label_password = tk.Label(root_create_account, font=font, text='Password:')
	label_password.place(rely=0.51, relwidth=0.25, relheight=0.25)

	# Entries
	entry_username = tk.Entry(root_create_account)
	entry_username.place(relx=0.25, rely=0.28, relwidth=0.25, relheight=0.17)

	entry_password = tk.Entry(root_create_account, state='disabled')
	entry_password.place(relx=0.25, rely=0.55, relwidth=0.25, relheight=0.17)

	label_password_not_working = tk.Label(root_create_account, text='Passwords comming soon')
	label_password_not_working.place(relx=0.18, rely=0.75, relwidth=0.35, relheight = 0.20)

	# Button
	button_create_account = tk.Button(root_create_account, text= 'Create\nAccount', command= lambda: add_user_to_dir(entry_username.get(), root_create_account))
	button_create_account.place(relx=0.7, rely=0.30, relwidth=0.2, relheight=0.4)


def open_catagory(catagory, username): # If catagory already exited, it just makes file_rel_path (a global variable) have the value of the name of the file to be opened. If its new, it creates the file
	global file_rel_path

	file_name = '%s_%s.xlsx' % (username, catagory)
	file_rel_path = os.path.join('users', username, catagory, file_name)
	print(file_rel_path, 'RELPATH')
	file_check_if_real = os.path.isfile(file_rel_path)
	print(file_check_if_real, 'CHECKING')
	if file_check_if_real == True:
		print('File Found')
	else:
		new_workbook = openpyxl.Workbook()
		database_sheet = new_workbook.active
		database_sheet.title = 'Database'
		new_workbook.create_sheet('Cost')
		cost_sheet = new_workbook['Cost']

		#Writes stuff in the new excel file
		database_sheet.cell(row=1, column=1, value='%s' % file_name)
		database_sheet.cell(row=2, column=1, value='Date Planted:')
		database_sheet.cell(row=3, column=1, value='Status:')
		database_sheet.cell(row=4, column=1, value='Entry Color:')
		database_sheet.cell(row=5, column=1, value='Seed Amount:')
		database_sheet.cell(row=6, column=1, value='Weight (g):')
		database_sheet.cell(row=7, column=1, value='Floating Seeds:')
		database_sheet.cell(row=8, column=1, value='Rinses per Day:')
		database_sheet.cell(row=9, column=1, value='Days in Dark:')
		database_sheet.cell(row=10, column=1, value='Total Days:')
		database_sheet.cell(row=11, column=1, value='Water Used (mL):')
		database_sheet.cell(row=12, column=1, value='Final Weight:')
		database_sheet.cell(row=13, column=1, value='Weight Growth:')
		database_sheet.cell(row=14, column=1, value='Water cost per L:')
		database_sheet.cell(row=15, column=1, value='Seed cost per tbsp:')
		database_sheet.cell(row=16, column=1, value='Water Cost Total:')
		database_sheet.cell(row=17, column=1, value='Seed Cost Total:')
		database_sheet.cell(row=18, column=1, value='Total Cost::')
		database_sheet.cell(row=19, column=1, value='Seed Supplier:')
		database_sheet.cell(row=21, column=1, value='Waterings:')

		cost_sheet.cell(row=2, column=1, value='Water Cost (L):')

		cost_sheet.cell(row=4, column=1, value='Seed Supplier')
		cost_sheet.cell(row=5, column=1, value='Seed Cost (tbsp):')

		cost_sheet.cell(row=2, column=2, value=0.20)
		cost_sheet.cell(row=5, column=2, value=0.20)

		new_workbook.save(file_rel_path)

	root_catagory.destroy()


	print(catagory)
	print(username)

# MAIN CODE START HERE

# Login Window
root_acc = tk.Tk()
root_acc.title('Log In to Sprouter')
root_acc.iconbitmap('broccoli.ico')

canvas = tk.Canvas(root_acc, width=600, height=400)
canvas.pack()

cwd = os.getcwd()
print(cwd, 'HEY')
print("WHAT")
cwd_dir_list = os.listdir()
print(len(cwd_dir_list))

users_file_found = False

for file in cwd_dir_list:
	if file == 'users':
		users_file_found = True

if users_file_found == False:
	os.mkdir('users')

user_list = os.listdir('users')
account_ammount = len(user_list)
print(account_ammount)


frame_acc_names = tk.Frame(root_acc, bg='pink')
frame_acc_names.place(relwidth=0.7, relheight=1)

frame_create_acc = tk.Frame(root_acc, bg='green')
frame_create_acc.place(relx=0.7, relwidth=0.3, relheight=1)


if account_ammount >= 1:
	button_acc_1 = tk.Button(frame_acc_names, text=user_list[0], command=lambda: login(user_list[0]))
	button_acc_1.place(relwidth=1, relheight=0.2)

if account_ammount >= 2:
	button_acc_2 = tk.Button(frame_acc_names, text=user_list[1], command=lambda: login(user_list[1]))
	button_acc_2.place(rely=0.2, relwidth=1, relheight=0.2)

if account_ammount >= 3:
	button_acc_3 = tk.Button(frame_acc_names, text=user_list[2], command=lambda: login(user_list[2]))
	button_acc_3.place(rely=0.4, relwidth=1, relheight=0.2)

if account_ammount >= 4:
	button_acc_4 = tk.Button(frame_acc_names, text=user_list[3], command=lambda: login(user_list[3]))
	button_acc_4.place(rely=0.6, relwidth=1, relheight=0.2)

if account_ammount >= 5:
	button_acc_5 = tk.Button(frame_acc_names, text=user_list[4], command=lambda: login(user_list[4]))
	button_acc_5.place(rely=0.8, relwidth=1, relheight=0.2)

if account_ammount > 5: # Current version only allows five accounts. Message desplays if more accounts are detected. Later will improve to allow to have more accounts and just be able to scroll through them
	root_more_than_five_accounts = tk.Tk()
	label_more_than_five_accounts = tk.Label(root_more_than_five_accounts, text='MORE THAN FIVE ACCOUNTS IN FILE.\nCURRENT VERSION DOES NOT ALLOW MORE THAN FIVE ACCOUNTS\n\nPlease remove from the folder if you wish to see all accounts')
	label_more_than_five_accounts.pack()

if account_ammount < 5:
	button_create_account = tk.Button(frame_create_acc, text='CREATE\nACCOUNT', command= create_account)
	button_create_account.place(relwidth=1, relheight=1)
else:
	button_create_account = tk.Button(frame_create_acc, text='CREATE\nACCOUNT', state= 'disabled', command= create_account)
	button_create_account.place(relwidth=1, relheight=1)

root_acc.mainloop()

# MAIN CODE START HERE
read_excel_var = read_excel() # This returns a list of the jars and the integer of amount of jars. This will be changed. The function will now take in a paramter. That parameter is the name of the username and subusername or file it is supposed to open. It might send username and subusername as one string or might send them as two seperate strings


jar_list = read_excel_var[0] # This separates the list from read_excel_var to its own variable
num_of_jars = read_excel_var[1]

screen_water_cost = float(read_excel_var[2])
#screen_seed_cost = float(read_excel_var[3])


root = tk.Tk()


root.title('Sprouter')
root.iconbitmap('broccoli.ico')
root.attributes('-fullscreen', True)
root.bind('<Escape>', exit_full_screen)


jar_dates = []
for i in range(len(jar_list)):
	date = jar_list[i][0].split('/')
	for j in range(len(date)):
		date[j] = int(date[j])

	date = datetime.date(date[2], date[0], date[1])
	jar_dates.append(date)





WIDTH = 800
HEIGHT = 600

image_background = tk.PhotoImage(file = 'farm3.png')
image_logo = tk.PhotoImage(file = 'sprouter_logo.png')


canvas = tk.Canvas(root, height=HEIGHT, width=WIDTH)
canvas.pack()

label_background = tk.Label(root, image=image_background)
label_background.place(relwidth=1, relheight=1)

label_logo = tk.Label(root, bg='#FCFBA6', image=image_logo)
label_logo.place(relx=0.2, rely=0.05, relwidth=0.3, relheight=0.15)

label_user_catagory = tk.Label(root, text=file_rel_path)
label_user_catagory.place(relx=0.4, rely=0.95, relwidth=0.2, relheight=0.05)

# FRAME FOR MAIN BUTTONS
frame_1 = tk.Frame(root, bg='pink')
frame_1.place(relx = 0.1, rely=0.5, relwidth=0.8, relheight=0.3)

# FRAMES FOR LEFT AND RIGHT SCROLLER BUTTONS
frame_left = tk.Frame(root)
frame_left.place(relx=0.03, rely=0.63, relwidth=0.04, relheight=0.06)

frame_right = tk.Frame(root)
frame_right.place(relx=0.93, rely=0.63, relwidth=0.04, relheight=0.06)

# FRAME FOR CREATE JAR
frame_create_jar = tk.Frame(root)
frame_create_jar.place(relx=0.7333, rely=0.3, relwidth=(0.5/3), relheight=0.08)

# FRAME FOR EXIT BUTTON

frame_exit = tk.Frame(root, bg='red')
frame_exit.place(relx=0.895, rely=0.91, relwidth=0.1, relheight=0.08)

# FRAME FOR PRICE BUTTONS

frame_price = tk.Frame(root, bg='red')
frame_price.place(relx=0.005, rely=0.91, relwidth=0.1, relheight=0.08)

# FRAME FOR SHOW ALL
frame_show_all = tk.Frame(root)
frame_show_all.place(relx=0.1, rely=0.3, relwidth=(0.5/3), relheight=0.08)


# MAIN BUTTONS
# NEED TO ADD FONT AND TEXT TO ALL BUTTONS
# Many if-statemets ahead to determine how many buttons to show

position = 0 # Helps determining where each button gets its information. Every time an arrow is clicked, this changes value
button_font = tk.font.Font(family='Helvetica', size=15, weight='bold')

# Activates buttons only if there are enough jars in the database.
if num_of_jars >= 5:
	button1 = tk.Button(frame_1, bg=jar_list[position+4][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+4][0],jar_list[position+4][1],(datetime.date.today()-jar_dates[position+4]).days,jar_list[position+4][8]), command=lambda: show_info(jar_list[position+4], position+4)) # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg=jar_list[position+3][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+3][0],jar_list[position+3][1],(datetime.date.today()-jar_dates[position+3]).days,jar_list[position+3][8]), command=lambda: show_info(jar_list[position+3], position+3))
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
if num_of_jars == 4:
	button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg=jar_list[position+3][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+3][0],jar_list[position+3][1],(datetime.date.today()-jar_dates[position+3]).days,jar_list[position+3][8]), command=lambda: show_info(jar_list[position+3], position+3))
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
if num_of_jars == 3:
	button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg='gray', state='disabled')
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg=jar_list[position+2][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+2][0],jar_list[position+2][1],(datetime.date.today()-jar_dates[position+2]).days,jar_list[position+2][8]), command=lambda: show_info(jar_list[position+2], position+2))
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
if num_of_jars == 2:
	button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg='gray', state='disabled')
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg='gray', state='disabled')
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg=jar_list[position+1][2],font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position+1][0],jar_list[position+1][1],(datetime.date.today()-jar_dates[position+1]).days,jar_list[position+1][8]), command=lambda: show_info(jar_list[position+1], position+1))
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
if num_of_jars == 1:
	button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg='gray', state='disabled')
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg='gray', state='disabled')
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg='gray', state='disabled')
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg=jar_list[position][2], font= button_font, text='%s\n\n%s\n\nDays:%s/%s' % (jar_list[position][0],jar_list[position][1],(datetime.date.today()-jar_dates[position]).days,jar_list[position][8]), command=lambda: show_info(jar_list[position], position)) # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)
if num_of_jars == 0:
	button1 = tk.Button(frame_1, bg='gray', state='disabled') # Leftmost button
	button1.place(relx=0.0277, rely=0.1, relwidth=0.166, relheight=0.8)

	button2 = tk.Button(frame_1, bg='gray', state='disabled')
	button2.place(relx=0.2214, rely=0.1, relwidth=0.166, relheight=0.8)

	button3 = tk.Button(frame_1, bg='gray', state='disabled')
	button3.place(relx=0.4151, rely=0.1, relwidth=0.166, relheight=0.8)

	button4 = tk.Button(frame_1, bg='gray', state='disabled')
	button4.place(relx=0.6088, rely=0.1, relwidth=0.166, relheight=0.8)

	button5 = tk.Button(frame_1, bg='gray', state='disabled') # Rightmost button
	button5.place(relx=0.8025, rely=0.1, relwidth=0.166, relheight=0.8)


# LEFT AND RIGHT BUTTONS

if num_of_jars >= 6:
	button_left = tk.Button(frame_left, bg='#FF7000',text='<<', command=move_left)
	button_left.place(relwidth=1, relheight=1)
else:
	button_left = tk.Button(frame_left, bg='gray',text='<<', command=move_left, state='disabled')
	button_left.place(relwidth=1, relheight=1)


button_right = tk.Button(frame_right, bg='gray', text='>>', command=move_right, state ='disabled')
button_right.place(relwidth=1, relheight=1)

# All mainmenu buttons will be shown here

# CREATE JAR BUTTON
button_create_jar = tk.Button(frame_create_jar, bg='sky blue', text='Create Jar', command=lambda: create_jar(num_of_jars))
button_create_jar.place(relwidth=1, relheight=1)

# EXIT BUTTON
button_exit = tk.Button(frame_exit, bg='#FE7A74', font='Calibri 14 bold', text='EXIT PROGRAM', command=root.quit)
button_exit.place(relwidth=1, relheight=1)

# PRICE BUTTON
button_price = tk.Button(frame_price, bg='#FE7A74', font='Calibri 14 bold', text='Water: $%.2f/L\n\n%s' % (screen_water_cost, 'Change Costs'), command=lambda: change_price(screen_water_cost))
button_price.place(relwidth=1, relheight=1)

# SHOW ALL BUTTON
button_show_all = tk.Button(frame_show_all, bg='sky blue', text='Show All', command=lambda: show_all())
button_show_all.place(relwidth=1, relheight=1)

print('heyy')
print(jar_list)
root.mainloop()
print('urmomisgay')


#### NEED TO WORK ON ####

#Make functioning button in 'harvested' windows that shows picture of jar


# Adding buttons to show jar entries in heeps of 20 or 50, instead of 5 at a time
# Make it so if someone who doestn alrady have an excel file called 'sprouts_tracker, it creates it for them


### Add 'What time are you watering this jar?' option to allow jars to be watered at different times.
### App sends you reminders when you need to water your plant. (I dont know how I would do this though unless app is running 24/7)

#### Change program so the date isnt the main identifier for the jar.
#### Right now it is not likely that I will have multiple jars in one day so identifying a jar by its date works. But if this becomes an industrial-size project (or just like a small business), I would plant multiple jars per day
#### I could fix this by having each jar randomly generate a serial code or maybe just a number starting from 0 and going up to infinity (each new jar always just takes the next number)

##### Counting 'floating seeds' might be unimportant. If I decide it is, I will remove it

###### Add error exceptions so program doesnt crash if user doesnt do everything as expected. Some places to add it would be:
###### - If the user closes the app before choosing a user or catagory, there is an error when the program continues and tries to open up an excel file since no excel file was ever chosen. error exception could just make it do nothing or print out (program closed) or something

####### Add seed provider catagory
####### Already added way to create jar with seed_provider attatched to it (meaning it writes it down in the database with all the other info)
####### What I need to do is 1. Add a button to create providers (this adds providers to the 'cost' sheet in excel and also allows user to type how much seeds cost with that provider, which will be written in the 'cost' sheet as well)
####### 2. When user 'harvests' a jar, price used to calculate the seed cost should be the correct vendor.
####### To do this, when clicking harvest, program should read from 'Database' sheet what vendor that jar was from, then go to the 'Cost' sheet, find that vendors name, and look right under for the price. That number (the price) will be used to calculate price
####### Also gotta add vendors name whenever you click a growing, ready, and harvested jar.


########## BEFORE PROGRAM WOULD GET COSTS OF WATER AND SEEDS WHEN USER HARBESTED. NOW THE PRICES ARE ASSIGNED WHEN CREATING THE JAR.
########## THIS MEANS I HAVE TO CHANGE CODE. AREAS OF CODE THAT NEED CHANGE: harvest, ready, growing (gotta change so harvest no longer assings prices (because they are already made) and that you can see the prices in the ready and growing stages).
########## also the "change price" button is now broken

################ FIX THE CHANGE PRICE BUTTON. MAKE IT SO YOU CAN CHANGE THE WATER PRICE OR THE PRICE OF A CERTAIN SEED PROVIDER. ALSO MAKE IT WHERE YOU CAN DELETE ONE SEED PROVIDER IF NEEDED ################